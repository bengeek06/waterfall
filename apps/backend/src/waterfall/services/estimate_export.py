"""Excel export for a project estimate: cost grid, subtotals and aggregates."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from waterfall.models.ms_core import MsProject
from waterfall.models.resources import Estimate, EstimateCostLine, EstimateLine
from waterfall.services.estimate_calculation import (
    EstimateAggregates,
    calculate_estimate_aggregates,
)

HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=13)


def _write_header(sheet: Worksheet, project: MsProject, estimate: Estimate) -> None:
    sheet["A1"] = f"Devis — {project.name}"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = "Version"
    sheet["B2"] = f"V{estimate.version_number} ({estimate.kind})"
    sheet["A3"] = "Statut"
    sheet["B3"] = estimate.status
    sheet["A4"] = "Devise"
    sheet["B4"] = estimate.currency_code
    sheet["A5"] = "Créé le"
    sheet["B5"] = estimate.created_at.strftime("%Y-%m-%d %H:%M")
    sheet["A6"] = "Validé le"
    sheet["B6"] = estimate.validated_at.strftime("%Y-%m-%d %H:%M") if estimate.validated_at else "-"


def _write_grid_sheet(
    sheet: Worksheet,
    project: MsProject,
    estimate: Estimate,
    labor_lines: list[EstimateLine],
    cost_lines: list[EstimateCostLine],
) -> None:
    _write_header(sheet, project, estimate)

    row = 8
    headers = ["Nature", "Catégorie", "Libellé", "Quantité", "Coût unitaire / horaire", "Montant"]
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=title)
        cell.font = HEADER_FONT
    row += 1

    total_labor = 0
    for line in labor_lines:
        sheet.cell(row=row, column=1, value="MO")
        sheet.cell(row=row, column=2, value=line.cost_category_code)
        sheet.cell(row=row, column=3, value=f"{line.task_name} ({line.role_name})")
        sheet.cell(row=row, column=4, value=float(line.hours))
        sheet.cell(row=row, column=5, value=float(line.hourly_rate))
        sheet.cell(row=row, column=6, value=float(line.budget_cost))
        total_labor += float(line.budget_cost)
        row += 1

    total_purchase = 0
    for cost_line in cost_lines:
        sheet.cell(row=row, column=1, value=cost_line.cost_type_code)
        sheet.cell(row=row, column=2, value=cost_line.cost_category_code)
        sheet.cell(row=row, column=3, value=cost_line.label)
        sheet.cell(row=row, column=4, value=float(cost_line.quantity))
        sheet.cell(row=row, column=5, value=float(cost_line.unit_cost))
        sheet.cell(row=row, column=6, value=float(cost_line.purchase_cost))
        total_purchase += float(cost_line.purchase_cost)
        row += 1

    row += 1
    for label, value in (
        ("Sous-total MO", total_labor),
        ("Sous-total Achat", total_purchase),
        ("PRU non chargé", total_labor + total_purchase),
    ):
        sheet.cell(row=row, column=5, value=label).font = HEADER_FONT
        sheet.cell(row=row, column=6, value=value).font = HEADER_FONT
        row += 1

    for column, width in enumerate([10, 16, 40, 12, 20, 14], start=1):
        sheet.column_dimensions[chr(64 + column)].width = width


def _write_aggregates_sheet(sheet: Worksheet, aggregates: EstimateAggregates) -> None:
    sheet["A1"] = "Agrégats"
    sheet["A1"].font = TITLE_FONT

    sheet["A3"] = "Total MO"
    sheet["B3"] = float(aggregates["total_labor_cost"])
    sheet["A4"] = "Total Achat"
    sheet["B4"] = float(aggregates["total_purchase_cost"])
    sheet["A5"] = "Total PRU non chargé"
    sheet["B5"] = float(aggregates["total_unburdened_cost"])

    row = 7
    sheet.cell(row=row, column=1, value="Catégorie").font = HEADER_FONT
    sheet.cell(row=row, column=2, value="Montant").font = HEADER_FONT
    row += 1
    for category_code, amount in aggregates["by_category"].items():
        sheet.cell(row=row, column=1, value=category_code)
        sheet.cell(row=row, column=2, value=float(amount))
        row += 1

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 16


def build_estimate_workbook(db: Session, project: MsProject, estimate: Estimate) -> bytes:
    """Build the devis Excel workbook: cost grid, subtotals and category aggregates."""
    labor_lines = (
        db.query(EstimateLine)
        .filter(EstimateLine.estimate_id == estimate.id)
        .filter(EstimateLine.role_id.isnot(None))
        .order_by(EstimateLine.id)
        .all()
    )
    cost_lines = (
        db.query(EstimateCostLine)
        .filter(EstimateCostLine.estimate_id == estimate.id)
        .order_by(EstimateCostLine.id)
        .all()
    )
    aggregates = calculate_estimate_aggregates(db, estimate.id)

    workbook = Workbook()
    grid_sheet = workbook.active
    assert grid_sheet is not None
    grid_sheet.title = "Devis"
    _write_grid_sheet(grid_sheet, project, estimate, labor_lines, cost_lines)

    aggregates_sheet = workbook.create_sheet("Agrégats")
    _write_aggregates_sheet(aggregates_sheet, aggregates)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
